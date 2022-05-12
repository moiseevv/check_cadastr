# This is a sample Python script.
from openpyxl import load_workbook as lwb
import openpyxl as op

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print_hi('PyCharm')

    name_source = 'final_filtr2.xlsx'
    file_source = lwb(name_source)
    ws = file_source.active
    kol =0

    wb_result = op.Workbook()
    ws_result = wb_result.active

    for i in range(1,ws.max_row+1):
        test_cadadastr = str(ws.cell(i,6).value)
        if test_cadadastr[0:3]=='54:':
            kol+=1
            print(test_cadadastr," ", kol)

            for u in range(1,ws.max_column+1):
                ws_result.cell(kol,u).value = ws.cell(i,u).value
wb_result.save('Only_54.xlsx')








# See PyCharm help at https://www.jetbrains.com/help/pycharm/
